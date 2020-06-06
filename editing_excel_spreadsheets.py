import openpyxl

wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'].value == None

sheet['A1'] = 42
sheet['A2'] = 'Hello'

wb.save('example1.xlsx')

sheet2 = wb.create_sheet()
wb.get_sheet_names()

sheet2.title = 'New Sheet'

wb.create_sheet(index=2, title='Other Sheet')