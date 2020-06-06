import openpyxl

workbook = openpyxl.load_workbook('example.xlsx')
workbook.sheetnames
sheet = workbook.get_sheet_by_name('Sheet1')
cell = sheet['A1']
str(cell.value)
str(sheet['A1'].value)
sheet['B1'].value
sheet['C1'].value

sheet.cell(row=1, column=2).value

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)